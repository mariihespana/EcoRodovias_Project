# Projeto Nova Dutra
# Objetivo: Coletar os dados de tráfego das pistas do site da Web e armazená-los em um arquivo excel estruturado para fins de análise

# Escrito por Mariana España Feitosa

# Para execução do script, instale o Anaconda, o python (e os módulos abaixo usando 'pip install (nome do modulo)') e o chromedriver
# Para execução do script, deixe o executável do chrome ('chromedriver.exe') na mesma pasta de execução do script.
# Se o seu Excel está em inglês, talvez você precise alterar a linha 137 deste código para 'Pasta1' ou conforme o Excel nomeia automaticamente as pastas ao ser criado.

# coding=<UTF-8>
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import openpyxl
import openpyxl.worksheet.worksheet
from openpyxl import load_workbook
import time
import os

class novaDutra():

    def __init__(self):
        # Comandos para abrir o Chrome
        chrome_options = Options()
        chrome_options.add_argument('--lang=pt-BR')
        self.driver = webdriver.Chrome(executable_path=r'./chromedriver.exe',options = chrome_options)

    def Iniciar(self):
        # Esta definição roda o programa quando já foi executado alguma vez, por isso, não precisa criar a planilha Excel
        self.acessar_site()
        self.info_window()
        self.coleta_dados()
        self.driver.close()
        self.armazenar_dados_em_planilha()
        print('Dados adicionados com sucesso!')

    def Iniciar_Primeira_Vez(self):
        # Esta definição roda o programa quando está rodando pela primeira vez o script e não há planilha criada para armazenamento dos dados
        self.acessar_site()
        self.info_window()
        self.coleta_dados()
        self.driver.close()
        self.criar_planilha()
        self.armazenar_dados_em_planilha()
        print('Dados adicionados com sucesso!')

    def acessar_site(self):
        # Acessa o site da Nova Dutra
        self.driver.get("http://www.novadutra.com.br/")

    def info_window(self):
        # Fecha a janela inicial do site
        time.sleep(3)
        close_window = self.driver.find_element_by_xpath('//*[@id="modal"]/div/header/a')
        close_window.click()

    def coleta_dados(self):
        # Coleta as informações do site
        dutra_elements = self.driver.find_elements_by_xpath('//*[contains(text(), "Presidente Dutra:")]')
        self.avenida = []
        self.sentido = []
        self.trafego = []
        self.pista = []
        self.obs = []
        self.kminicial = []
        self.kmfinal = []
        self.data = []
        self.hora = []

        for element in dutra_elements:
            # Transforma a informação coletada em texto
            texto = element.text
            # Quebra o texto em pedaços
            lista = texto.split()
            # Coleta apenas o pedaço do texto desejado, usando o separador espaço " "
            avenida = ' '.join(map(str,lista[0:2]))
            self.avenida.append(avenida[:-1])
            # print(self.avenida)
            self.sentido.append(' '.join(map(str,lista[4:7])))
            # print(self.sentido)
            self.trafego.append(lista[lista.index('tráfego')+1])
            # print(self.trafego)
            pista = ' '.join(map(str,lista[lista.index('pista')+1:lista.index('pista')+2]))
            self.pista.append(pista[:-1])
            # print(self.pista)
            obs = ' '.join(map(str,lista[lista.index('Obs:')+1:lista.index('km')])).replace(".","",-1)
            self.obs.append(obs)
            # print(self.obs)
            self.kminicial.append(lista[lista.index('inicial:')+1])
            # print(self.kminicial)
            self.kmfinal.append(lista[lista.index('final:')+1])
            # print(self.kmfinal)
            data = lista[lista.index('final:')+2]
            self.data.append(data)
            # print(self.data)
            self.hora.append(lista[lista.index('final:')+3])
            # print(self.hora)

        #     add_dados.append(element.text)
        # print(add_dados)

        # # Para cada elemento encontrado, adicionar a uma lista
        # for element in dutra_elements:
        #     print(element.text)
        #     add_dados.append(element.text)
        # print(add_dados)

        # # Salvar dados no arquivo dados.txt
        # with open('dados.txt', 'a', encoding='utf-8') as arquivo:
        #     for item in add_dados:
        #         arquivo.write(item + '\n')


    def criar_planilha(self):
        # Cria o arquivo Excel
        self.planilha = openpyxl.Workbook()
        # Cria a aba 'Dados Pistas'
        self.planilha.create_sheet('Dados Pistas')
        self.planilha_valores = self.planilha['Dados Pistas']
        # Insere os valores do cabeçalho
        self.planilha_valores['A1'].value = 'Avenida'
        self.planilha_valores['B1'].value = 'Sentido'
        self.planilha_valores['C1'].value = 'Tráfego'
        self.planilha_valores['D1'].value = 'Pista'
        self.planilha_valores['E1'].value = 'Obs.'
        self.planilha_valores['F1'].value = 'Km inicial'
        self.planilha_valores['G1'].value = 'Km final'
        self.planilha_valores['H1'].value = 'Data'
        self.planilha_valores['I1'].value = 'Hora'
        # Salva a planilha
        self.planilha.save('Dados Concorrencia.xlsx')
        self.sheet = self.planilha['Dados Pistas']

    def armazenar_dados_em_planilha(self):
        # Armazena os dados no arquivo Excel
        file = os.getcwd() + os.sep + '\Dados Concorrencia.xlsx'
        file = openpyxl.load_workbook(filename=file)
        # Deleta a aba 'Sheet' criada automaticamente no arquivo Excel
        if 'Sheet' in file.sheetnames:
            del file['Sheet']
        # print(file['Dados Pistas'].max_row)
        novos_dados = []
        # Para cada dado coletado no site, inclui as informações nas colunas conforme a ordem criada.
        for indice in range(0,len(self.avenida)):
            nova_linha = [self.avenida[indice],
                            self.sentido[indice],
                            self.trafego[indice],
                            self.pista[indice],
                            self.obs[indice],
                            self.kminicial[indice],
                            self.kmfinal[indice],
                            self.data[indice],
                            self.hora[indice]]
            # Insere linha
            # openpyxl.worksheet.worksheet.Worksheet.insert_rows(file['Dados Pistas'], 2)
            # Inclui a nova linha dentro da aba 'Dados Pistas'
            file['Dados Pistas'].append(nova_linha)
        # print(file['Dados Pistas'].max_row)
        # Salva o arquivo
        file.save('Dados Concorrencia.xlsx')

# Cria uma instância para a classe novaDutra
dados_concorrencia = novaDutra()

# Condição, se o arquivo Excel já existe, rode o programa sem criar um arquivo novo
if os.path.exists(os.getcwd() + os.sep + '\Dados Concorrencia.xlsx') == True:
    dados_concorrencia.Iniciar()
# Condição, se o arquivo Excel NÃO existe, roda o programa e cria um arquivo novo
if os.path.exists(os.getcwd() + os.sep + '\Dados Concorrencia.xlsx') == False:
    dados_concorrencia.Iniciar_Primeira_Vez()
